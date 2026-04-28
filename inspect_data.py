"""Phase 2 data inspection for Gonzalez-Sulser GRIN2B share.

Run with the gonzalez-sulser-lab-to-nwb-env conda env:
    conda activate gonzalez-sulser-lab-to-nwb-env
    python inspect_data.py
"""
from __future__ import annotations

import json
import struct
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

DATA_ROOT = Path("H:/Gonzalez-Sulser-CN-data-share")
EEG_DIR = DATA_ROOT / "Chronic EEG recordings"
SLEEP_DIR = DATA_ROOT / "Sleep state classifications"
SEIZURE_DIR = DATA_ROOT / "Seizure timestamps"
LIGHT_DIR = DATA_ROOT / "Light cycle timing metadata"

REPORT: dict = {}


# ---------------------------------------------------------------------------
# 1. TAINI .dat binary layout probes
# ---------------------------------------------------------------------------
def probe_dat_layout():
    """Test plausible (n_channels, dtype, sample_rate) combos against file size.

    Goal: confirm n_channels=16, dtype=int16, fs in {250, 250.4, 256, 19531.25}.
    """
    files = sorted(EEG_DIR.glob("*.dat"))
    out = {"n_files": len(files), "files": []}
    for fp in files:
        size = fp.stat().st_size
        candidates = []
        for n_ch in (16, 14, 32):
            for dt_size, dt_name in ((2, "int16"), (4, "float32"), (4, "int32")):
                if size % (n_ch * dt_size) != 0:
                    continue
                n_samples = size // (n_ch * dt_size)
                for fs in (250.0, 250.4, 256.0, 500.0, 1000.0, 19531.25):
                    dur_s = n_samples / fs
                    dur_d = dur_s / 86400
                    if 0.5 < dur_d < 14:
                        candidates.append(
                            dict(
                                n_ch=n_ch,
                                dtype=dt_name,
                                fs=fs,
                                n_samples=n_samples,
                                dur_days=round(dur_d, 3),
                            )
                        )
        out["files"].append(dict(name=fp.name, size_bytes=size, candidates=candidates[:8]))
    REPORT["dat_layout_probe"] = out
    print(f"[dat] inspected {len(files)} .dat files; first file candidates:")
    for c in out["files"][0]["candidates"]:
        print("   ", c)


# ---------------------------------------------------------------------------
# 2. Read first samples assuming int16 / 16ch interleaved and inspect ranges
# ---------------------------------------------------------------------------
def probe_dat_values(n_ch: int = 16, n_samples: int = 100_000):
    """Read a small chunk and report per-channel statistics."""
    fp = sorted(EEG_DIR.glob("*.dat"))[0]
    with open(fp, "rb") as f:
        raw = f.read(n_ch * 2 * n_samples)
    arr = np.frombuffer(raw, dtype="<i2").reshape(-1, n_ch)
    stats = []
    for ch in range(n_ch):
        x = arr[:, ch].astype(np.int64)
        stats.append(
            dict(
                ch=ch,
                min=int(x.min()),
                max=int(x.max()),
                mean=round(float(x.mean()), 3),
                std=round(float(x.std()), 3),
                n_unique=int(np.unique(x).size),
            )
        )
    REPORT["dat_first_chunk_stats"] = dict(file=fp.name, n_samples_read=arr.shape[0], stats=stats)
    print(f"[dat] {fp.name} — first {arr.shape[0]} samples per ch:")
    for s in stats:
        print(
            f"   ch{s['ch']:02d}  range=[{s['min']:>7},{s['max']:>7}]  mean={s['mean']:>9}  std={s['std']:>9}  uniq={s['n_unique']}"
        )


# ---------------------------------------------------------------------------
# 3. Sleep state CSV
# ---------------------------------------------------------------------------
def probe_sleep_csvs():
    """Inspect dge_ok.csv label distribution, pw_spectrum, and Channels.csv head."""
    subj = SLEEP_DIR / "GRIN2B_129"
    out = {}

    # dge_ok
    f = subj / "GRIN2B_129_BL1-dge_ok.csv"
    df = pd.read_csv(f)
    out["dge_ok"] = dict(
        path=str(f),
        rows=len(df),
        columns=list(df.columns),
        head=df.head(10).to_dict(orient="records"),
        unique_values=df.iloc[:, 0].value_counts().head(20).to_dict(),
    )
    print(f"[sleep] dge_ok rows={len(df)} cols={list(df.columns)}")
    print(f"   value counts (top 20): {out['dge_ok']['unique_values']}")

    # pw_spectrum
    f = subj / "GRIN2B_129_BL1-pw_spectrum.csv"
    df = pd.read_csv(f)
    out["pw_spectrum"] = dict(
        path=str(f),
        rows=len(df),
        columns=list(df.columns),
        head=df.head(5).to_dict(orient="records"),
        tail=df.tail(3).to_dict(orient="records"),
        first_col_unique_top=df.iloc[:, 0].astype(str).value_counts().head(10).to_dict(),
    )
    print(f"[sleep] pw_spectrum rows={len(df)} cols={list(df.columns)}")
    print(f"   head: {out['pw_spectrum']['head'][:3]}")
    print(f"   tail: {out['pw_spectrum']['tail']}")

    # Channels.csv: huge — read in chunks for stats
    f = subj / "GRIN2B_129_BL1_Channels.csv"
    head = pd.read_csv(f, nrows=20)
    out["channels_head"] = dict(
        path=str(f), columns=list(head.columns), head=head.to_dict(orient="records")
    )
    # streaming stats
    n_rows = 0
    sums = np.zeros(2, dtype=np.float64)
    sumsq = np.zeros(2, dtype=np.float64)
    mins = np.full(2, np.inf)
    maxs = np.full(2, -np.inf)
    n_nonzero = np.zeros(2, dtype=np.int64)
    first_nonzero_row = [None, None]
    for chunk in pd.read_csv(f, chunksize=1_000_000):
        a = chunk.to_numpy(dtype=np.float64)
        n_rows += a.shape[0]
        sums += a.sum(axis=0)
        sumsq += (a * a).sum(axis=0)
        mins = np.minimum(mins, a.min(axis=0))
        maxs = np.maximum(maxs, a.max(axis=0))
        for c in range(2):
            nz = a[:, c] != 0
            n_nonzero[c] += int(nz.sum())
            if first_nonzero_row[c] is None and nz.any():
                first_nonzero_row[c] = int(np.argmax(nz)) + (n_rows - a.shape[0])
    means = sums / n_rows
    stds = np.sqrt(sumsq / n_rows - means * means)
    out["channels_stats"] = dict(
        n_rows=n_rows,
        n_rows_eq_24h_at_250p4=round(n_rows / 250.4, 2),
        per_col=[
            dict(
                col=int(c),
                min=float(mins[c]),
                max=float(maxs[c]),
                mean=round(float(means[c]), 3),
                std=round(float(stds[c]), 3),
                n_nonzero=int(n_nonzero[c]),
                first_nonzero_row=first_nonzero_row[c],
            )
            for c in range(2)
        ],
    )
    print(f"[sleep] Channels.csv n_rows={n_rows}  (= {round(n_rows/250.4,2)} s @ 250.4 Hz)")
    for s in out["channels_stats"]["per_col"]:
        print(
            f"   col{s['col']}  range=[{s['min']:.1f},{s['max']:.1f}]  mean={s['mean']:.2f}  std={s['std']:.2f}  nonzero={s['n_nonzero']}  first_nz_row={s['first_nonzero_row']}"
        )

    REPORT["sleep_csvs"] = out


# ---------------------------------------------------------------------------
# 4. Seizure CSVs
# ---------------------------------------------------------------------------
def probe_seizure_csvs():
    files = sorted(SEIZURE_DIR.glob("*.csv"))
    out = {"n_files": len(files), "samples": []}
    for f in files[:5]:
        try:
            df = pd.read_csv(f, sep="\t")
            if df.shape[1] == 1:
                df = pd.read_csv(f)
        except Exception as e:
            out["samples"].append(dict(file=f.name, error=str(e)))
            continue
        out["samples"].append(
            dict(
                file=f.name,
                rows=len(df),
                cols=list(df.columns),
                head=df.head(3).to_dict(orient="records"),
                max_sec_end=float(df["sec_end"].max()) if "sec_end" in df.columns else None,
            )
        )
    print(f"[seizure] n_files={len(files)}; samples:")
    for s in out["samples"]:
        print(f"   {s['file']}: rows={s.get('rows')} cols={s.get('cols')} max_end={s.get('max_sec_end')}")
    REPORT["seizure_csvs"] = out


# ---------------------------------------------------------------------------
# 5. Seiz subfolder (DGE_SWDs, Seiz_Totals, Seizures inside seiz/)
# ---------------------------------------------------------------------------
def probe_seiz_subfolder():
    subj = SLEEP_DIR / "GRIN2B_129" / "seiz"
    out = {"files": []}
    for f in sorted(subj.glob("*.csv")):
        df = pd.read_csv(f)
        out["files"].append(
            dict(
                name=f.name,
                rows=len(df),
                cols=list(df.columns),
                head=df.head(3).to_dict(orient="records"),
            )
        )
    # compare top-level Seizures.csv with seiz/Seizures.csv for same animal
    top = SEIZURE_DIR / "GRIN2B_129_BL1_Seizures.csv"
    inner = subj / "GRIN2B_129_BL1_Seizures.csv"
    if top.exists() and inner.exists():
        a = pd.read_csv(top, sep="\t")
        b = pd.read_csv(inner, sep="\t")
        if a.shape[1] == 1:
            a = pd.read_csv(top)
        if b.shape[1] == 1:
            b = pd.read_csv(inner)
        out["dup_check"] = dict(
            top_shape=list(a.shape),
            inner_shape=list(b.shape),
            equal=bool(a.equals(b)) if a.shape == b.shape else False,
        )
    print("[seiz subfolder] files:")
    for s in out["files"]:
        print(f"   {s['name']}: rows={s['rows']} cols={s['cols']}")
    if "dup_check" in out:
        print(f"[seiz subfolder] dup check: {out['dup_check']}")
    REPORT["seiz_subfolder"] = out


# ---------------------------------------------------------------------------
# 6. Light cycle xlsx
# ---------------------------------------------------------------------------
def probe_light_cycle_xlsx():
    f = LIGHT_DIR / "Sample_start_end_GRIN2B.xlsx"
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    out = {"path": str(f), "sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        out["sheets"].append(
            dict(
                name=sheet_name,
                n_rows=len(rows),
                n_cols=max((len(r) for r in rows), default=0),
                first_10=[list(r) for r in rows[:10]],
                last_3=[list(r) for r in rows[-3:]],
            )
        )
    print(f"[xlsx] {f.name}; sheets:")
    for s in out["sheets"]:
        print(f"   {s['name']}: {s['n_rows']} rows × {s['n_cols']} cols")
        for r in s["first_10"]:
            print(f"     {r}")
    REPORT["light_cycle_xlsx"] = out


# ---------------------------------------------------------------------------
# 7. Cross-check: does Channels.csv col0 == raw .dat single channel?
# ---------------------------------------------------------------------------
def cross_check_channels_vs_dat():
    """Compare Channels.csv col0 against each of the 16 channels of the raw .dat,
    over the first 10 minutes of BL1.

    The Sample_start_end xlsx tells us BL1 sample offset; we need the offset.
    Without it, brute-force search for the alignment.
    """
    # Identify the .dat file matching subject 129 BL1
    dat = EEG_DIR / "TAINI_1047_B_Grin2B_129_Baseline-2021_03_26-0000.dat"
    if not dat.exists():
        print("[xcheck] dat for 129 not found")
        return

    chan = SLEEP_DIR / "GRIN2B_129" / "GRIN2B_129_BL1_Channels.csv"

    # Read first 60 seconds of Channels.csv (60 * 250.4 ~ 15024 rows) — but its
    # head was zero. Read a chunk far enough in that there's signal.
    df = pd.read_csv(chan, skiprows=range(1, 250_400 * 60 + 1), nrows=2500)  # at minute 60
    print("[xcheck] Channels.csv chunk @ ~60min:", df.shape, df.head(3).to_dict(orient="records"))
    if df.empty or df.iloc[:, 0].abs().sum() == 0:
        print("   chunk all-zero — Channels.csv may not be raw data")
        REPORT["cross_check"] = dict(status="channels_csv_all_zero_in_chunk")
        return

    target = df.iloc[:, 0].to_numpy(dtype=np.float64)
    # Read corresponding chunk from .dat at sample offset = 60*60*250.4 = 901,440
    n_ch = 16
    sample_offset = int(60 * 60 * 250.4)
    n_samples = len(target)
    # Try a window around this offset for alignment
    out = {"target_stats": dict(min=float(target.min()), max=float(target.max())), "matches": []}
    with open(dat, "rb") as f:
        for off in (sample_offset - 1000, sample_offset, sample_offset + 1000):
            f.seek(off * n_ch * 2)
            raw = f.read(n_samples * n_ch * 2)
            arr = np.frombuffer(raw, dtype="<i2").reshape(-1, n_ch)
            for ch in range(n_ch):
                x = arr[:, ch].astype(np.float64)
                # compare correlation / equality after possible scaling
                if x.std() == 0:
                    continue
                corr = float(np.corrcoef(x, target)[0, 1])
                if abs(corr) > 0.5:
                    out["matches"].append(dict(off=off, ch=ch, corr=round(corr, 4)))
    print(f"[xcheck] matches (|corr|>0.5):")
    for m in out["matches"]:
        print(f"   {m}")
    REPORT["cross_check"] = out


# ---------------------------------------------------------------------------
# 8. dge_ok value <-> Channels duration check
# ---------------------------------------------------------------------------
def probe_epoch_alignment():
    """17,280 epochs × 5 s = 86,400 s = 24 h. Channels.csv rows ÷ 250.4 = 86,395.X.
    Compute exact ratio and report.
    """
    n_chan_rows = REPORT.get("sleep_csvs", {}).get("channels_stats", {}).get("n_rows")
    if n_chan_rows:
        for fs in (250.0, 250.4, 256.0):
            secs = n_chan_rows / fs
            print(f"[align] Channels n_rows / {fs} Hz = {secs:.2f} s ({secs/3600:.4f} h, {secs/5:.2f} epochs of 5s)")
        REPORT["alignment"] = dict(channels_csv_rows=n_chan_rows)


def main():
    print("=" * 70)
    probe_dat_layout()
    print("=" * 70)
    probe_dat_values()
    print("=" * 70)
    probe_sleep_csvs()
    print("=" * 70)
    probe_seizure_csvs()
    print("=" * 70)
    probe_seiz_subfolder()
    print("=" * 70)
    probe_light_cycle_xlsx()
    print("=" * 70)
    probe_epoch_alignment()
    print("=" * 70)
    cross_check_channels_vs_dat()

    out_path = Path(__file__).parent / "inspection_report.json"
    with open(out_path, "w") as f:
        json.dump(REPORT, f, indent=2, default=str)
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
